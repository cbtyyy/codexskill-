from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


SKILL_ROOT = Path(__file__).resolve().parents[1]
LIBRARY_ROOT = SKILL_ROOT / "assets" / "dual_factory_library"
FACE_NAMES = ("top", "front", "right")
WORKBOOK_BACKGROUND = (255, 245, 212)
EXPECTED_VERSIONS = {
    "table1": "table1_three_visible_faces_v4",
    "table2": "table2_three_visible_faces_v4",
}
EXPECTED_TEXTURE_PROFILE = {
    "name": "fine_print_v4_balanced_smooth",
    "source_profile": "fine_print_v3",
    "unsharp_radius": 0.9,
    "unsharp_percent": 145,
    "unsharp_threshold": 1,
    "contrast_factor": 1.09,
    "saturation_factor": 1.018,
    "soften_radius": 0.9,
    "soften_blend": 0.3,
    "texture_interpolation": "Linear",
    "geometry_policy": "Preserve source artwork, line positions, and face registration exactly.",
}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def validate_manifest(library_id: str) -> dict:
    manifest_path = LIBRARY_ROOT / f"{library_id}_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest["library_id"] != library_id:
        raise RuntimeError(f"Wrong library ID in {manifest_path}")
    if manifest.get("version") != EXPECTED_VERSIONS[library_id]:
        raise RuntimeError(f"Wrong texture-library version in {manifest_path}")
    if manifest.get("texture_profile") != EXPECTED_TEXTURE_PROFILE:
        raise RuntimeError(f"Wrong texture profile in {manifest_path}")

    expected_prefix = "表1-" if library_id == "table1" else "表2-"
    seen = set()
    face_count = 0
    workbook_thumbnail_count = 0
    for record in manifest["records"]:
        key = record["key"]
        if key in seen:
            raise RuntimeError(f"Duplicate key: {key}")
        seen.add(key)
        if not key.startswith(f"{library_id}:"):
            raise RuntimeError(f"Invalid namespace: {key}")
        if record["detail_label"] != f"{expected_prefix}{record['id']}":
            raise RuntimeError(f"Invalid detail label: {key}")
        for relative_path in record.get("workbook_thumbnail_paths", []):
            thumbnail_path = LIBRARY_ROOT / relative_path
            with Image.open(thumbnail_path) as thumbnail:
                thumbnail.load()
                rgb = thumbnail.convert("RGB")
                corners = (
                    rgb.getpixel((0, 0)),
                    rgb.getpixel((rgb.width - 1, 0)),
                    rgb.getpixel((0, rgb.height - 1)),
                    rgb.getpixel((rgb.width - 1, rgb.height - 1)),
                )
                if any(pixel != WORKBOOK_BACKGROUND for pixel in corners):
                    raise RuntimeError(f"Thumbnail background mismatch: {key}")
            workbook_thumbnail_count += 1
        if not record.get("scene_eligible"):
            continue
        if set(record["files"]) != set(FACE_NAMES):
            raise RuntimeError(f"Incomplete faces: {key}")
        for face_name in FACE_NAMES:
            file_info = record["files"][face_name]
            path = LIBRARY_ROOT / file_info["path"]
            if sha256(path) != file_info["sha256"]:
                raise RuntimeError(f"Hash mismatch: {key} {face_name}")
            with Image.open(path) as image:
                image.load()
                if image.size != (512, 512) or image.mode != "RGB":
                    raise RuntimeError(f"Invalid face image: {key} {face_name}")
            face_count += 1
    return {
        "records": len(manifest["records"]),
        "scene_eligible": sum(bool(r.get("scene_eligible")) for r in manifest["records"]),
        "face_references": face_count,
        "workbook_thumbnails": workbook_thumbnail_count,
        "texture_profile": manifest["texture_profile"],
    }


def validate_workbook(filename: str, expected_records: int) -> dict:
    path = LIBRARY_ROOT / "workbooks" / filename
    workbook = load_workbook(path, read_only=False, data_only=False)
    sheet = workbook["颗粒"]
    if any(cell.value == "产品名称" for row in sheet.iter_rows() for cell in row):
        raise RuntimeError(f"Product-name field found in {filename}")
    if (sheet["A1"].value, sheet["B1"].value, sheet["C1"].value) != (
        "编号",
        "图片",
        "尺寸",
    ):
        raise RuntimeError(f"Unexpected headers in {filename}")
    record_count = sum(
        1
        for row in sheet.iter_rows()
        for cell in row
        if cell.column % 3 == 1 and cell.row >= 2 and cell.value not in (None, "")
    )
    if record_count != expected_records:
        raise RuntimeError(f"Record count mismatch in {filename}")
    return {"records": record_count, "images": len(sheet._images)}


def main() -> None:
    table1 = validate_manifest("table1")
    table2 = validate_manifest("table2")
    table1["workbook"] = validate_workbook("表格1.xlsx", table1["records"])
    table2["workbook"] = validate_workbook("表格2.xlsx", table2["records"])
    print(json.dumps({"status": "pass", "table1": table1, "table2": table2}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
