from __future__ import annotations

import hashlib
import json
import os
import random
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageChops, ImageDraw, ImageEnhance, ImageFilter, ImageFont


SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))

import factory_texture_redraw_rules as legacy


SOURCE_XLSX = os.environ.get("MAGNETIC_FACTORY_XLSX")
SOURCE_SKU_SHEET = os.environ.get("MAGNETIC_FACTORY_SKU_SHEET")
FACE_DIR = SKILL_ROOT / "assets/locked_factory_faces_v2"
MANIFEST_PATH = SKILL_ROOT / "references/locked-factory-faces-v2.json"
CONTACT_PATH = SKILL_ROOT / "assets/locked-factory-face-contact-sheet-v2.png"

SKU_IDS = (
    1, 2, 6, 9, 10, 11, 13, 14, 15, 16, 17, 18, 19, 20, 22, 23,
    25, 26, 27, 32, 33, 35, 45, 48, 57, 59, 60, 66, 67, 71, 73,
    74, 75, 76, 77, 78, 79, 80, 81, 82, 83, 86, 87, 101, 102,
    109, 129, 134, 145, 150, 151, 153, 154, 155, 156, 157, 161,
    166, 176, 180, 181, 182, 184, 185,
)

NATURAL_SKUS = {
    1, 2, 6, 9, 10, 11, 13, 14, 15, 16, 17, 18, 19, 22, 23, 27,
    48, 57, 81, 86, 87, 145, 150, 151, 153, 154, 156, 157, 161,
    176, 180, 181, 185,
}
SUBTLE_GRAIN_SKUS = {
    25, 26, 32, 33, 35, 45, 59, 60, 66, 67, 71, 73, 74, 75,
    76, 77, 78, 79, 80, 82, 83, 155, 166, 184,
}
PALE_NATURAL_SKUS = {145, 150, 180, 181}


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    name = "msyhbd.ttc" if bold else "msyh.ttc"
    return ImageFont.truetype(f"C:/Windows/Fonts/{name}", size)


def rgb(hex_value: str) -> tuple[int, int, int]:
    value = hex_value.lstrip("#")
    return tuple(int(value[index:index + 2], 16) for index in (0, 2, 4))


def solid(color: str | tuple[int, int, int]) -> Image.Image:
    return Image.new("RGB", (16, 16), rgb(color) if isinstance(color, str) else color)


def dirt(seed: int = 1) -> Image.Image:
    return legacy.noise_texture(
        (126, 88, 58),
        [(92, 62, 45), (151, 105, 67), (179, 130, 78), (105, 73, 50)],
        seed,
        52,
        2,
    )


def grass(side: bool = False) -> Image.Image:
    if not side:
        return legacy.noise_texture(
            (87, 168, 64),
            [(52, 128, 48), (112, 190, 72), (143, 204, 77), (71, 147, 55)],
            202,
            58,
            2,
        )
    image = dirt(203)
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, 15, 3), fill=(90, 174, 64))
    for x, depth in ((1, 5), (4, 4), (7, 6), (11, 4), (14, 5)):
        draw.rectangle((x, 3, min(15, x + 1), depth), fill=(60, 133, 49))
    return image


def netherrack(seed: int) -> Image.Image:
    return legacy.noise_texture(
        (111, 52, 49),
        [(79, 35, 39), (144, 65, 55), (172, 77, 61), (91, 42, 46)],
        seed,
        56,
        2,
    )


def birch(top: bool = False) -> Image.Image:
    if top:
        image = solid((218, 203, 156))
        draw = ImageDraw.Draw(image)
        for inset, color in ((1, (93, 79, 58)), (4, (168, 145, 101)), (7, (104, 86, 59))):
            draw.rectangle((inset, inset, 15 - inset, 15 - inset), outline=color)
        return image
    image = solid((226, 222, 201))
    draw = ImageDraw.Draw(image)
    for x, y, w in ((1, 2, 4), (9, 1, 3), (4, 6, 5), (11, 8, 4), (0, 12, 4), (7, 14, 3)):
        draw.rectangle((x, y, min(15, x + w), min(15, y + 1)), fill=(68, 66, 61))
    return image


def pumpkin(top: bool = False) -> Image.Image:
    image = solid((207, 104, 34))
    draw = ImageDraw.Draw(image)
    if top:
        draw.rectangle((6, 6, 9, 9), fill=(89, 92, 45))
        for inset in (1, 4):
            draw.rectangle((inset, inset, 15 - inset, 15 - inset), outline=(171, 79, 29))
    else:
        for x in (3, 7, 11):
            draw.line((x, 0, x, 15), fill=(159, 73, 29))
        draw.rectangle((3, 5, 5, 7), fill=(68, 48, 30))
        draw.rectangle((10, 5, 12, 7), fill=(68, 48, 30))
        draw.polygon(((4, 10), (7, 12), (11, 9), (12, 12), (4, 13)), fill=(246, 187, 48))
    return image


def dark_planks() -> Image.Image:
    return legacy.planks((83, 59, 48))


def redstone_lamp() -> Image.Image:
    image = solid((71, 57, 43))
    draw = ImageDraw.Draw(image)
    draw.rectangle((1, 1, 14, 14), outline=(40, 37, 32), width=2)
    for polygon in (
        ((7, 1), (11, 4), (8, 7), (4, 4)),
        ((1, 7), (4, 4), (7, 8), (4, 11)),
        ((14, 7), (11, 4), (8, 8), (11, 11)),
        ((7, 14), (4, 11), (8, 8), (11, 11)),
    ):
        draw.polygon(polygon, fill=(231, 168, 93), outline=(116, 76, 45))
    draw.rectangle((6, 6, 9, 9), fill=(255, 210, 117))
    return image


def barrel(top: bool = False) -> Image.Image:
    if top:
        image = solid((152, 103, 59))
        draw = ImageDraw.Draw(image)
        draw.rectangle((2, 2, 13, 13), outline=(64, 45, 32), width=2)
        draw.line((3, 7, 12, 7), fill=(95, 61, 38))
        return image
    image = legacy.planks((139, 91, 52), vertical=True)
    draw = ImageDraw.Draw(image)
    for y in (2, 12):
        draw.rectangle((0, y, 15, y + 1), fill=(64, 49, 40))
    return image


def honey() -> Image.Image:
    return legacy.noise_texture(
        (226, 147, 39),
        [(255, 190, 60), (187, 108, 31), (250, 215, 98)],
        8100,
        38,
        2,
    )


def hay(top: bool = False) -> Image.Image:
    image = solid((190, 157, 54))
    draw = ImageDraw.Draw(image)
    if top:
        for inset in (2, 5):
            draw.rectangle((inset, inset, 15 - inset, 15 - inset), outline=(132, 99, 35))
    else:
        for x in range(1, 16, 3):
            draw.line((x, 0, x, 15), fill=(218, 186, 70))
        draw.rectangle((0, 6, 15, 8), fill=(143, 73, 43))
    return image


def snow_top() -> Image.Image:
    image = solid((243, 245, 242))
    draw = ImageDraw.Draw(image)
    for x, y in ((2, 3), (10, 2), (6, 8), (13, 11), (3, 13)):
        draw.point((x, y), fill=(215, 225, 228))
    return image


def snow_stone(side_kind: str = "stone") -> dict[str, Image.Image]:
    side = legacy.cobble(1810, (137, 140, 140)) if side_kind == "cobble" else legacy.stone_texture(1450, (136, 138, 138))
    front = side.copy()
    for image in (side, front):
        draw = ImageDraw.Draw(image)
        draw.rectangle((0, 0, 15, 3), fill=(240, 243, 240))
        for x, depth in ((1, 5), (5, 4), (9, 6), (13, 4)):
            draw.rectangle((x, 3, min(15, x + 1), depth), fill=(220, 226, 226))
    return {"top": snow_top(), "side": side, "front": front}


def snow_grass() -> dict[str, Image.Image]:
    side = grass(side=True)
    front = side.copy()
    for image in (side, front):
        draw = ImageDraw.Draw(image)
        draw.rectangle((0, 0, 15, 3), fill=(241, 245, 241))
        for x, depth in ((2, 5), (7, 4), (12, 6)):
            draw.rectangle((x, 3, min(15, x + 1), depth), fill=(218, 226, 225))
    return {"top": snow_top(), "side": side, "front": front}


def glass() -> Image.Image:
    image = solid((196, 229, 236))
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, 15, 15), outline=(87, 105, 106), width=2)
    draw.line((3, 13, 13, 3), fill=(238, 251, 251), width=2)
    draw.line((1, 10, 7, 4), fill=(152, 207, 219))
    return image


def dense_leaves(seed: int = 1560) -> Image.Image:
    image = legacy.noise_texture(
        (54, 121, 49),
        [(32, 90, 39), (76, 151, 59), (111, 172, 71), (45, 108, 47)],
        seed,
        72,
        2,
    )
    draw = ImageDraw.Draw(image)
    for x, y in ((2, 4), (7, 2), (12, 5), (4, 10), (10, 12), (14, 14)):
        draw.line((x, y, min(15, x + 2), max(0, y - 2)), fill=(138, 193, 86))
    return image


def crate(top: bool = False) -> Image.Image:
    image = legacy.planks((139, 83, 49))
    draw = ImageDraw.Draw(image)
    draw.rectangle((1, 1, 14, 14), outline=(77, 48, 34), width=2)
    if top:
        draw.line((2, 2, 13, 13), fill=(184, 116, 64), width=2)
        draw.line((13, 2, 2, 13), fill=(184, 116, 64), width=2)
    else:
        draw.line((2, 3, 13, 12), fill=(183, 112, 61), width=2)
        draw.line((13, 3, 2, 12), fill=(183, 112, 61), width=2)
    return image


def ice(seed: int = 1760) -> Image.Image:
    image = legacy.noise_texture(
        (103, 190, 232),
        [(62, 154, 216), (147, 220, 245), (82, 176, 226), (196, 239, 250)],
        seed,
        42,
        2,
    )
    draw = ImageDraw.Draw(image)
    draw.line((1, 12, 6, 7, 10, 9, 15, 3), fill=(222, 248, 252))
    return image


def snowflake() -> Image.Image:
    image = ice(1800)
    draw = ImageDraw.Draw(image)
    color = (233, 249, 252)
    draw.line((8, 2, 8, 13), fill=color)
    draw.line((2, 8, 13, 8), fill=color)
    draw.line((4, 4, 12, 12), fill=color)
    draw.line((12, 4, 4, 12), fill=color)
    return image


def snowman_faces() -> dict[str, Image.Image]:
    top = solid((245, 247, 243))
    face = solid((239, 246, 246))
    draw = ImageDraw.Draw(face)
    draw.rectangle((2, 5, 4, 7), fill=(56, 101, 152))
    draw.rectangle((11, 5, 13, 7), fill=(56, 101, 152))
    draw.point((3, 6), fill=(238, 242, 242))
    draw.point((12, 6), fill=(238, 242, 242))
    draw.polygon(((7, 7), (12, 9), (7, 10)), fill=(231, 73, 48))
    draw.arc((4, 8, 12, 14), 15, 165, fill=(42, 46, 49), width=1)
    draw.rectangle((0, 0, 15, 2), fill=(48, 151, 189))
    draw.rectangle((2, 1, 10, 3), fill=(57, 173, 204))
    profile = solid((239, 246, 246))
    pd = ImageDraw.Draw(profile)
    pd.rectangle((0, 0, 15, 2), fill=(48, 151, 189))
    pd.polygon(((5, 7), (12, 9), (5, 10)), fill=(231, 73, 48))
    return {"top": top, "side": face, "front": profile}


def terracotta(seed: int = 1850) -> Image.Image:
    return legacy.noise_texture(
        (164, 91, 61),
        [(128, 67, 48), (193, 116, 74), (145, 78, 55)],
        seed,
        24,
        2,
    )


def custom_faces(sku: int) -> dict[str, Image.Image]:
    if sku == 1:
        return {name: dirt(100 + index) for index, name in enumerate(("top", "side", "front"))}
    if sku == 2:
        return {"top": grass(), "side": grass(True), "front": grass(True)}
    if sku == 22:
        return {name: netherrack(2200 + index) for index, name in enumerate(("top", "side", "front"))}
    if sku == 25:
        return {"top": birch(True), "side": birch(), "front": birch()}
    if sku == 32:
        return {"top": pumpkin(True), "side": pumpkin(), "front": pumpkin()}
    if sku == 35:
        return {name: dark_planks() for name in ("top", "side", "front")}
    if sku == 71:
        return {name: redstone_lamp() for name in ("top", "side", "front")}
    if sku == 77:
        return {"top": barrel(True), "side": barrel(), "front": barrel()}
    if sku == 81:
        return {name: honey() for name in ("top", "side", "front")}
    if sku == 82:
        return {"top": hay(True), "side": hay(), "front": hay()}
    if sku == 145:
        return snow_stone("stone")
    if sku == 150:
        return snow_grass()
    if sku == 155:
        return {name: glass() for name in ("top", "side", "front")}
    if sku == 156:
        return {name: dense_leaves(1560 + index) for index, name in enumerate(("top", "side", "front"))}
    if sku == 166:
        return {"top": crate(True), "side": crate(), "front": crate()}
    if sku == 176:
        return {name: ice(1760 + index) for index, name in enumerate(("top", "side", "front"))}
    if sku == 180:
        return {name: snowflake() for name in ("top", "side", "front")}
    if sku == 181:
        return snow_stone("cobble")
    if sku == 182:
        return snowman_faces()
    if sku == 185:
        return {name: terracotta(1850 + index) for index, name in enumerate(("top", "side", "front"))}
    raise KeyError(sku)


def draw_faces(sku: int) -> tuple[str, dict[str, Image.Image]]:
    try:
        return "manual_redraw_existing", legacy.make_faces(sku)
    except KeyError:
        return "manual_redraw_extended", custom_faces(sku)


def grain_layer(seed: int, size: int, amplitude: int, chroma: int) -> Image.Image:
    rng = random.Random(seed)
    pixels = []
    for _ in range(size * size):
        value = rng.randint(-amplitude, amplitude)
        pixels.append((
            max(0, min(255, 128 + value + rng.randint(-chroma, chroma))),
            max(0, min(255, 128 + value + rng.randint(-chroma, chroma))),
            max(0, min(255, 128 + value + rng.randint(-chroma, chroma))),
        ))
    layer = Image.new("RGB", (size, size))
    layer.putdata(pixels)
    # A 64-cell print grain is four times denser than the base 16-cell drawing.
    # Nearest expansion keeps PNGs compact; at catalog scale each grain cell is
    # subpixel and reads as surface texture rather than another visible mosaic.
    return layer.resize((512, 512), Image.Resampling.NEAREST)


def add_grain(image: Image.Image, sku: int, face_name: str, amplitude: int, chroma: int) -> Image.Image:
    face_index = {"top": 1, "side": 2, "front": 3}[face_name]
    grain = grain_layer(sku * 1009 + face_index * 97, 64, amplitude, chroma)
    return ImageChops.add(image, grain, scale=1.0, offset=-128)


def finish_face(image: Image.Image, sku: int, face_name: str) -> Image.Image:
    source = image.convert("RGB")
    if sku in NATURAL_SKUS:
        # Factory natural prints retain their block identity but contain much
        # finer ink variation than a literal 16x16 nearest-neighbor upscale.
        crisp = source.resize((512, 512), Image.Resampling.NEAREST)
        soft = source.resize((512, 512), Image.Resampling.BICUBIC)
        output = Image.blend(crisp, soft, 0.30)
        amplitude = 2 if sku in PALE_NATURAL_SKUS else 6
        chroma = 1 if sku in PALE_NATURAL_SKUS else 3
        output = add_grain(output, sku, face_name, amplitude, chroma)
        output = output.filter(ImageFilter.UnsharpMask(radius=0.9, percent=34, threshold=3))
    else:
        # Geometric prints keep exact lines; only a restrained print-grain pass
        # is allowed on wood, brick, props, and other non-character surfaces.
        output = source.resize((512, 512), Image.Resampling.NEAREST)
        output = output.filter(ImageFilter.GaussianBlur(0.75))
        if sku in SUBTLE_GRAIN_SKUS:
            output = add_grain(output, sku, face_name, 3, 1)
        output = output.filter(ImageFilter.UnsharpMask(radius=0.8, percent=38, threshold=3))
    return ImageEnhance.Color(output).enhance(1.045)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def source_preview(sheet: Image.Image, sku: int) -> Image.Image:
    cell_w = sheet.width / 15
    cell_h = sheet.height / 13
    index = sku - 1
    column = index % 15
    row = index // 15
    # The upper portion is the factory cube; the bottom contains its number.
    box = (
        round(column * cell_w),
        round(row * cell_h),
        round((column + 1) * cell_w),
        round(row * cell_h + cell_h * 0.78),
    )
    return sheet.crop(box).convert("RGBA")


def read_excel_skus(path: Path) -> dict[int, dict[str, str | None]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    records: dict[int, dict[str, str]] = {}
    for id_column in range(1, worksheet.max_column + 1, 4):
        for row in range(2, worksheet.max_row + 1):
            raw_id = worksheet.cell(row=row, column=id_column).value
            code = str(raw_id).strip() if raw_id is not None else ""
            if not code.isdigit():
                continue
            sku = int(code)
            name = str(worksheet.cell(row=row, column=id_column + 1).value or "").strip()
            size = str(worksheet.cell(row=row, column=id_column + 3).value or "").strip()
            records[sku] = {
                "code": code.zfill(3),
                "name": name or None,
                "size": size or None,
                "metadata_status": "complete" if name and size else "id_only",
            }
    workbook.close()
    return records


def main() -> None:
    if not SOURCE_XLSX or not SOURCE_SKU_SHEET:
        raise RuntimeError(
            "Set MAGNETIC_FACTORY_XLSX and MAGNETIC_FACTORY_SKU_SHEET before rebuilding."
        )
    source_xlsx = Path(SOURCE_XLSX)
    source_sku_sheet = Path(SOURCE_SKU_SHEET)
    if not source_xlsx.exists() or not source_sku_sheet.exists():
        raise FileNotFoundError("Factory Excel and SKU overview are required")
    excel_skus = read_excel_skus(source_xlsx)
    missing_excel_skus = [sku for sku in SKU_IDS if sku not in excel_skus]
    if missing_excel_skus:
        raise ValueError(f"Locked SKUs missing from Excel cells: {missing_excel_skus}")
    FACE_DIR.mkdir(parents=True, exist_ok=True)
    for stale in FACE_DIR.glob("*.png"):
        stale.unlink()

    manifest = {
        "version": "locked_factory_faces_v2",
        "source_excel": {
            "name": "磁力颗粒编号.xlsx",
            "bundled": False,
            "role": "formal SKU cell IDs; product names and sizes where populated",
        },
        "source_sku_sheet": {
            "name": "磁力颗粒SKU大图.png",
            "bundled": False,
            "role": "factory color and pattern reference",
        },
        "mapping_policy": "Use Excel cell IDs; never map embedded images by drawing order.",
        "render_policy": (
            "Procedurally redrawn square faces with factory-calibrated multiscale print grain; "
            "factory overview crops are contact-sheet references and never become Blender materials."
        ),
        "face_size": 512,
        "skus": [],
    }
    previews: list[tuple[int, str, dict[str, Image.Image]]] = []
    for sku in SKU_IDS:
        method, faces = draw_faces(sku)
        finished = {
            name: finish_face(faces[name], sku, name)
            for name in ("top", "side", "front")
        }
        files = {}
        for name, image in finished.items():
            path = FACE_DIR / f"sku_{sku:03d}_{name}.png"
            image.save(path, optimize=True)
            files[name] = {
                "path": path.relative_to(SKILL_ROOT).as_posix(),
                "sha256": sha256(path),
            }
        manifest["skus"].append({
            "sku": excel_skus[sku]["code"],
            "name": excel_skus[sku]["name"],
            "size": excel_skus[sku]["size"],
            "metadata_status": excel_skus[sku]["metadata_status"],
            "method": method,
            "finish_profile": (
                "natural_multiscale"
                if sku in NATURAL_SKUS
                else "geometric_subtle_grain"
                if sku in SUBTLE_GRAIN_SKUS
                else "geometric_exact"
            ),
            "files": files,
        })
        previews.append((sku, method, finished))

    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    columns = 4
    tile_w, tile_h = 500, 178
    rows = (len(previews) + columns - 1) // columns
    sheet = Image.new("RGB", (columns * tile_w, rows * tile_h), (247, 247, 245))
    draw = ImageDraw.Draw(sheet)
    factory_sheet = Image.open(source_sku_sheet).convert("RGB")
    for index, (sku, method, faces) in enumerate(previews):
        x = index % columns * tile_w
        y = index // columns * tile_h
        source = source_preview(factory_sheet, sku)
        source.thumbnail((120, 120), Image.Resampling.LANCZOS)
        sheet.paste(source.convert("RGB"), (x + 8, y + 34), source.getchannel("A"))
        draw.text((x + 8, y + 6), f"SKU {sku:03d}  厂图→标准正方形重画", font=font(15, True), fill=(35, 38, 41))
        for face_index, name in enumerate(("top", "side", "front")):
            preview = faces[name].resize((110, 110), Image.Resampling.LANCZOS)
            px = x + 140 + face_index * 118
            sheet.paste(preview, (px, y + 34))
            draw.text((px, y + 148), name, font=font(13), fill=(75, 79, 82))
    sheet.save(CONTACT_PATH, quality=96)
    print(json.dumps({"faces": len(SKU_IDS) * 3, "manifest": str(MANIFEST_PATH), "contact": str(CONTACT_PATH)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
